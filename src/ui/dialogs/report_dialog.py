"""
Panther Assessment - Report Dialog Mixin
Report generation methods for MainWindow
"""

import re
import subprocess
import platform
import traceback
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMessageBox, QProgressDialog, QFileDialog
)
from PyQt6.QtCore import Qt

class ReportDialogMixin:
    """Mixin providing report generation for MainWindow"""

    def generate_report_from_outcomes(self, dialog):
        """Generate report with configured outcomes"""
        if self.outcome_list.count() == 0:
            QMessageBox.warning(
                dialog,
                "No Outcomes",
                "Please create at least one outcome before generating a report."
            )
            return
        
        dialog.accept()
        
        # Collect outcomes
        outcomes = []
        for i in range(self.outcome_list.count()):
            item = self.outcome_list.item(i)
            outcomes.append(item.data(Qt.ItemDataRole.UserRole))
        
        # Show progress
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        
        try:
            # Fetch student data and generate report
            self.fetch_data_and_generate_report(outcomes)
            
        except Exception as e:
            import traceback
            QApplication.restoreOverrideCursor()
            error_details = traceback.format_exc()
            QMessageBox.critical(
                self,
                "Error",
                f"Error generating report:\n{str(e)}"
            )
    
    def fetch_data_and_generate_report(self, outcomes):
        """Fetch student data and generate Excel report - OPTIMIZED"""
        import pandas as pd
        from datetime import datetime
        from PyQt6.QtWidgets import QProgressDialog, QFileDialog
        from PyQt6.QtCore import Qt
        
        # Create progress dialog
        progress = QProgressDialog("Generating report...", "Cancel", 0, 100, self)
        progress.setWindowTitle("Please Wait")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        # Get unique course IDs from assignments
        course_ids = set()
        all_assignment_ids = set()
        for outcome in outcomes:
            for assignment in outcome['assignments']:
                # Handle both course_ids (list) and course_id (single)
                assignment_course_ids = assignment.get('course_ids', [])
                if not assignment_course_ids:
                    # Fallback to singular course_id
                    single_id = assignment.get('course_id')
                    if single_id:
                        assignment_course_ids = [single_id]
                
                for cid in assignment_course_ids:
                    course_ids.add(cid)
                
                assignment_id = assignment.get('id')
                if assignment_id:
                    all_assignment_ids.add(assignment_id)
        
        # If course_id not in assignment, get from course_info
        if not course_ids:
            course_ids = set(c['id'] for c in self.course_info)
        
        progress.setLabelText("Fetching students...")
        progress.setValue(10)
        QApplication.processEvents()
        
        # Fetch all students from all courses
        all_students = {}
        student_courses = {}  # Track which courses each student is enrolled in: {student_id: [course_ids]}
        for course_id in course_ids:
            enrollments = self.canvas_client.get_enrollments(course_id)
            for enrollment in enrollments:
                user = enrollment.get('user', {})
                student_id = user.get('id')
                if student_id:
                    if student_id not in all_students:
                        all_students[student_id] = {
                            'id': student_id,
                            'sis_user_id': user.get('sis_user_id', ''),
                            'name': user.get('name', 'Unknown'),
                            'sortable_name': user.get('sortable_name', 'Unknown')
                        }
                    # Track all courses this student is enrolled in
                    if student_id not in student_courses:
                        student_courses[student_id] = []
                    student_courses[student_id].append(course_id)
        
        if not all_students:
            progress.close()
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(
                self,
                "No Students Found",
                "No students found in the selected courses.\n"
                "Make sure students are enrolled in these courses."
            )
            return
        
        progress.setLabelText("Fetching assignment submissions...")
        progress.setValue(20)
        QApplication.processEvents()
        
        # Fetch submissions - organized by student's actual enrolled course
        # Structure: {assignment_id: {student_id: score}}
        all_submissions = {}
        
        total_assignments = len(all_assignment_ids)
        for idx, outcome in enumerate(outcomes):
            for assignment in outcome['assignments']:
                assignment_id = assignment.get('id')  # This is just for the dict key
                course_ids_list = assignment.get('course_ids', [assignment.get('course_id')])
                assignment_ids_by_course = assignment.get('assignment_ids_by_course', {})
                
                if assignment_id not in all_submissions:
                    all_submissions[assignment_id] = {}
                    
                    # For each course this assignment appears in
                    for course_id in course_ids_list:
                        # CRITICAL: Use the correct assignment_id for THIS course
                        course_assignment_id = assignment_ids_by_course.get(course_id, assignment_id)
                        submissions = self.canvas_client.get_submissions(course_id, course_assignment_id)
                        
                        # Only store submissions for students enrolled in THIS specific course
                        for submission in submissions:
                            student_id = submission.get('user_id')
                            score = submission.get('score')
                            workflow_state = submission.get('workflow_state', '')
                            
                            is_enrolled = student_id in student_courses and course_id in student_courses[student_id]
                            
                            # Check: is this student enrolled in THIS course?
                            if is_enrolled:
                                # Only store graded submissions with actual scores
                                if workflow_state == 'graded' and score is not None:
                                    all_submissions[assignment_id][student_id] = score
            
            # Update progress
            current = 20 + int((idx / len(outcomes)) * 15)
            progress.setValue(current)
            progress.setLabelText(f"Fetching submissions... ({idx + 1}/{len(outcomes)} outcomes)")
            QApplication.processEvents()
            
            if progress.wasCanceled():
                QApplication.restoreOverrideCursor()
                return
        
        # ============ NEW: PRE-FETCH QUIZ DATA ============
        progress.setLabelText("Pre-fetching quiz data...")
        progress.setValue(35)
        QApplication.processEvents()
        
        all_quiz_data = {}  # {(course_id, quiz_id, student_id): {group_id: {'points': X, 'count': Y}}}
        
        quiz_assignments = []
        for outcome in outcomes:
            for assignment in outcome['assignments']:
                quiz_ids_by_course = assignment.get('quiz_ids_by_course', {})
                if quiz_ids_by_course:
                    quiz_assignments.append((assignment, quiz_ids_by_course))
        
        for idx, (assignment, quiz_ids_by_course) in enumerate(quiz_assignments):
            course_ids_list = assignment.get('course_ids', [])
            assignment_id = assignment.get('id')
            assignment_ids_by_course = assignment.get('assignment_ids_by_course', {})
            
            for course_id in course_ids_list:
                course_quiz_id = quiz_ids_by_course.get(course_id)
                if not course_quiz_id:
                    continue
                
                # Build question_id -> {group_id, name, points} mapping
                quiz_questions = self.canvas_client.get_quiz_questions(course_id, course_quiz_id)
                question_to_group = {}  # question_id -> str(group_id)
                question_meta = {}      # question_id -> {name, points_possible}
                for q in quiz_questions:
                    qid = q.get('id')
                    gid = q.get('quiz_group_id')
                    if qid and gid:
                        question_to_group[qid] = str(gid)
                        question_meta[qid] = {
                            'name': q.get('question_name') or q.get('question_text', f'Q{qid}')[:40],
                            'points_possible': q.get('points_possible', 0)
                        }
                
                # Get course-specific assignment ID
                course_assignment_id = assignment_ids_by_course.get(course_id, assignment_id)
                
                # Get all quiz submissions for this course
                quiz_subs = self.canvas_client.get_quiz_submissions(course_id, course_quiz_id)
                
                for quiz_sub in quiz_subs:
                    student_id = quiz_sub.get('user_id')
                    
                    # Only fetch for enrolled students
                    if student_id not in student_courses or course_id not in student_courses[student_id]:
                        continue
                    
                    # Fetch assignment submission with submission_history to get actual points
                    try:
                        response = self.canvas_client.session.get(
                            f"{self.canvas_client.base_url}/api/v1/courses/{course_id}/assignments/{course_assignment_id}/submissions/{student_id}",
                            params={'include[]': 'submission_history'},
                            timeout=30
                        )
                        
                        if response.status_code == 200:
                            sub_data = response.json()
                            submission_history = sub_data.get('submission_history', [])
                            
                            # Get submission_data from the most recent attempt
                            if submission_history:
                                latest = submission_history[-1]
                                submission_data = latest.get('submission_data', [])
                                
                                # Sum points by group
                                key = (course_id, course_quiz_id, student_id)
                                if key not in all_quiz_data:
                                    all_quiz_data[key] = {}
                                
                                for item in submission_data:
                                    if not isinstance(item, dict):
                                        continue
                                    question_id = item.get('question_id')
                                    points = item.get('points', 0) or 0
                                    
                                    group_id = question_to_group.get(question_id)
                                    if group_id:
                                        # Group-level totals
                                        if group_id not in all_quiz_data[key]:
                                            all_quiz_data[key][group_id] = {'points': 0, 'count': 0, 'questions': {}}
                                        all_quiz_data[key][group_id]['points'] += points
                                        all_quiz_data[key][group_id]['count'] += 1
                                        # Question-level detail
                                        qmeta = question_meta.get(question_id, {})
                                        all_quiz_data[key][group_id]['questions'][question_id] = {
                                            'points': points,
                                            'points_possible': qmeta.get('points_possible', 0),
                                            'name': qmeta.get('name', f'Q{question_id}')
                                        }
                    except Exception:
                        pass
                        pass        
            
            current = 35 + int((idx / len(quiz_assignments)) * 20) if quiz_assignments else 35
            progress.setValue(current)
            progress.setLabelText(f"Pre-fetching quiz data... ({idx + 1}/{len(quiz_assignments)} quizzes)")
            QApplication.processEvents()
        
        # ============ NEW: PRE-FETCH RUBRIC DATA ============
        progress.setLabelText("Pre-fetching rubric data...")
        progress.setValue(55)
        QApplication.processEvents()
        
        all_rubric_data = {}  # {(course_id, assignment_id, student_id): rubric_assessment}
        
        rubric_assignments = []
        for outcome in outcomes:
            parts_config = outcome.get('parts_config', {})
            for assignment in outcome['assignments']:
                assignment_id = assignment.get('id')
                parts_key = assignment_id if assignment_id in parts_config else str(assignment_id)
                if parts_key in parts_config:
                    selected_parts = parts_config[parts_key]
                    if any(p.get('type') == 'rubric_criterion' for p in selected_parts if isinstance(p, dict)):
                        rubric_assignments.append(assignment)
        
        for idx, assignment in enumerate(rubric_assignments):
            assignment_id = assignment.get('id')
            course_ids_list = assignment.get('course_ids', [])
            assignment_ids_by_course = assignment.get('assignment_ids_by_course', {})
            
            for course_id in course_ids_list:
                course_assignment_id = assignment_ids_by_course.get(course_id, assignment_id)
                submissions = self.canvas_client.get_submissions(course_id, course_assignment_id)
                
                for submission in submissions:
                    student_id = submission.get('user_id')
                    
                    # Only store for enrolled students
                    if student_id not in student_courses or course_id not in student_courses[student_id]:
                        continue
                    
                    rubric_assessment = submission.get('rubric_assessment', {})
                    if rubric_assessment:
                        key = (course_id, course_assignment_id, student_id)
                        all_rubric_data[key] = rubric_assessment
            
            current = 55 + int((idx / len(rubric_assignments)) * 15) if rubric_assignments else 55
            progress.setValue(current)
            progress.setLabelText(f"Pre-fetching rubric data... ({idx + 1}/{len(rubric_assignments)} assignments)")
            QApplication.processEvents()
        
        progress.setLabelText("Calculating outcome scores...")
        progress.setValue(70)
        QApplication.processEvents()
        
        NA_NO_ASSIGNMENT = "#N/A - Assignment Not Found"
        NA_NO_SCORE = "#N/A - No Score Data"
        
        # Build report rows — one per student
        total_pct_rows = []   # Tab 1: outcome totals
        raw_data_rows = []    # Tab 2: per-part scores

        for student_idx, (student_id, student) in enumerate(all_students.items()):
            if progress.wasCanceled():
                QApplication.restoreOverrideCursor()
                return

            student_info = {
                'Student ID': student.get('sis_user_id', ''),
                'Student Name': student['sortable_name'],
                'Course ID': student_courses.get(student_id, [None])[0]
            }

            total_row = dict(student_info)
            raw_row = dict(student_info)

            for outcome in outcomes:
                outcome_name = outcome['name']
                parts_config = outcome.get('parts_config', {})
                total_earned = 0
                total_possible = 0
                missing_type = None  # None, NA_NO_ASSIGNMENT, or NA_NO_SCORE

                for assignment in outcome['assignments']:
                    assignment_id = assignment.get('id')
                    assignment_name = assignment.get('name', 'Unknown')
                    assignment_possible = assignment.get('points_possible', 0)
                    course_ids_list = assignment.get('course_ids', [assignment.get('course_id')])
                    assignment_ids_by_course = assignment.get('assignment_ids_by_course', {})
                    quiz_ids_by_course = assignment.get('quiz_ids_by_course', {})

                    student_course_ids = student_courses.get(student_id, [])
                    relevant_courses = [cid for cid in course_ids_list if cid in student_course_ids]

                    # Check if assignment exists in student's section at all
                    assignment_found = len(relevant_courses) > 0

                    parts_key = assignment_id if assignment_id in parts_config else str(assignment_id)
                    if parts_key in parts_config:
                        selected_parts = parts_config[parts_key]
                        if not isinstance(selected_parts, list):
                            selected_parts = []

                        for part in selected_parts:
                            if not isinstance(part, dict):
                                continue
                            part_type = part.get('type')
                            part_name = part.get('description') or part.get('group_name') or 'Unknown'
                            col = f"{outcome_name} - {assignment_name} - {part_name}"

                            if not assignment_found:
                                raw_row[col] = NA_NO_ASSIGNMENT
                                if missing_type != NA_NO_ASSIGNMENT:
                                    missing_type = NA_NO_ASSIGNMENT
                                continue

                            if part_type == 'quiz_group':
                                group_ids_by_course = part.get('group_ids_by_course', {})
                                pick_count_by_course = part.get('pick_count_by_course', {})
                                question_points_by_course = part.get('question_points_by_course', {})
                                found = False
                                for course_id in relevant_courses:
                                    group_id = str(group_ids_by_course.get(str(course_id), ''))
                                    course_quiz_id = quiz_ids_by_course.get(course_id)
                                    if not group_id or not course_quiz_id:
                                        continue
                                    key = (course_id, course_quiz_id, student_id)
                                    if key in all_quiz_data and group_id in all_quiz_data[key]:
                                        gdata = all_quiz_data[key][group_id]
                                        earned = gdata.get('points', 0)
                                        count = gdata.get('count', 0)
                                        qpts = question_points_by_course.get(str(course_id), 0)
                                        possible = count * qpts
                                        raw_row[f"{col} (Group Total)"] = earned
                                        total_earned += earned
                                        total_possible += possible
                                        found = True
                                        # Add per-question columns using group's points_per_question
                                        for qid, qdata in gdata.get('questions', {}).items():
                                            qname = qdata.get('name', f'Q{qid}')
                                            qscore = qdata.get('points', 0)
                                            # Use group-defined points per question, not question metadata
                                            q_pts_possible = qpts
                                            qpct = round((qscore / q_pts_possible) * 100) if q_pts_possible else NA_NO_SCORE
                                            q_col = f"{col} - {qname}"
                                            raw_row[f"{q_col} (Raw)"] = qscore
                                            raw_row[f"{q_col} (%)"] = qpct
                                        break
                                if not found:
                                    raw_row[col] = NA_NO_SCORE
                                    if missing_type is None:
                                        missing_type = NA_NO_SCORE

                            elif part_type == 'rubric_criterion':
                                criterion_ids_by_course = part.get('criterion_ids_by_course', {})
                                criterion_points_possible = part.get('points', 0)
                                found = False
                                for course_id in relevant_courses:
                                    criterion_id = criterion_ids_by_course.get(str(course_id))
                                    if not criterion_id:
                                        continue
                                    course_assignment_id = assignment_ids_by_course.get(course_id, assignment_id)
                                    key = (course_id, course_assignment_id, student_id)
                                    if key in all_rubric_data:
                                        rubric_assessment = all_rubric_data[key]
                                        score = None
                                        for cid_key in [criterion_id, str(criterion_id)]:
                                            if cid_key in rubric_assessment:
                                                score = rubric_assessment[cid_key].get('points', 0)
                                                break
                                        if score is not None:
                                            raw_row[col] = score
                                            total_earned += score
                                            total_possible += criterion_points_possible
                                            found = True
                                            break
                                if not found:
                                    raw_row[col] = NA_NO_SCORE
                                    if missing_type is None:
                                        missing_type = NA_NO_SCORE
                    else:
                        # No parts — use full assignment score
                        col = f"{outcome_name} - {assignment_name}"
                        if not assignment_found:
                            raw_row[col] = NA_NO_ASSIGNMENT
                            if missing_type != NA_NO_ASSIGNMENT:
                                missing_type = NA_NO_ASSIGNMENT
                        else:
                            student_score = all_submissions.get(assignment_id, {}).get(student_id)
                            if student_score is not None:
                                raw_row[col] = student_score
                                total_earned += student_score
                                total_possible += assignment_possible
                            else:
                                raw_row[col] = NA_NO_SCORE
                                if missing_type is None:
                                    missing_type = NA_NO_SCORE

                # Total % tab: outcome name is the column header
                if missing_type is not None:
                    total_row[outcome_name] = missing_type
                elif total_possible > 0:
                    total_row[outcome_name] = round((total_earned / total_possible) * 100)
                else:
                    total_row[outcome_name] = NA_NO_SCORE

            total_pct_rows.append(total_row)
            raw_data_rows.append(raw_row)

            current = 70 + int((student_idx / len(all_students)) * 20)
            progress.setValue(current)
            QApplication.processEvents()

        progress.setLabelText("Creating Excel file...")
        progress.setValue(90)
        QApplication.processEvents()

        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        if not total_pct_rows:
            progress.close()
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(self, "No Data", "No student data found.")
            return

        # Sort by student name
        total_pct_rows.sort(key=lambda r: r.get('Student Name', ''))
        raw_data_rows.sort(key=lambda r: r.get('Student Name', ''))

        # Generate filename
        course_code = "Course"
        if self.course_info:
            first_course_name = self.course_info[0].get('name', '')
            import re
            match = re.match(r'^([A-Z]+\s*\d+)', first_course_name)
            if match:
                course_code = match.group(1).replace(' ', '')

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{course_code}_outcome_report_{timestamp}.xlsx"

        last_dir = getattr(self, '_last_report_dir', None)
        default_path = str(Path(last_dir) / default_filename) if last_dir else str(Path.home() / "Desktop" / default_filename)
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Save Report",
            default_path,
            "Excel Files (*.xlsx)"
        )

        if not filepath:
            progress.close()
            QApplication.restoreOverrideCursor()
            return

        filepath = Path(filepath)
        self._last_report_dir = str(filepath.parent)

        wb = openpyxl.Workbook()

        def write_sheet(wb, sheet_name, rows, first_sheet=False):
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)

            if not rows:
                return ws

            headers = list(rows[0].keys())
            header_fill = PatternFill(start_color="862633", end_color="862633", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

            na_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
            na_font = Font(color="C62828", italic=True)

            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    val = row_data.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if val in (NA_NO_ASSIGNMENT, NA_NO_SCORE):
                        cell.fill = na_fill
                        cell.font = na_font

            # Auto-size columns
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(header))
                for row_idx in range(2, len(rows) + 2):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            return ws

        write_sheet(wb, "Total %", total_pct_rows, first_sheet=True)
        write_sheet(wb, "Raw Data", raw_data_rows)

        wb.save(filepath)

        progress.setValue(100)
        progress.close()
        QApplication.restoreOverrideCursor()

        # Show success message
        result = QMessageBox.information(

            self,
            "Report Generated",
            f"Report generated successfully!\n\n"
            f"File: {filepath.name}\n"
            f"Location: {filepath.parent}\n"
            f"Students: {len(all_students)}\n"
            f"Outcomes: {len(outcomes)}\n\n"
            f"Would you like to open the report folder?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if result == QMessageBox.StandardButton.Yes:
            import subprocess
            import platform
            if platform.system() == 'Windows':
                subprocess.run(['explorer', str(filepath.parent)])
            elif platform.system() == 'Darwin':
                subprocess.run(['open', str(filepath.parent)])
            else:
                subprocess.run(['xdg-open', str(filepath.parent)])

